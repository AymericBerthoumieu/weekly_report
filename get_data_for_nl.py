import datetime as dt
import pandas as pd
import requests
from lxml import html
import numpy as np


class LoadDataWeekChange:

    def __init__(self, source: pd.Series, init: pd.Series, type_spot: pd.Series):
        self.source = source
        self.ytd_df = init
        self.type_spot = type_spot

        self.current = pd.DataFrame()
        self.previous = pd.DataFrame()
        self.change = None

    def load_data(self):
        for name in self.source.index:
            print(f'Loading {name}...')
            self.current[name], self.previous[name] = self.parser(name, self.source.loc[name])
        self.current = self.current.applymap(
            lambda x: x if type(x) == float else float(x) if not '%' in x else float(x[:-1])).T
        self.previous = self.previous.applymap(
            lambda x: x if type(x) == float else float(x) if not '%' in x else float(x[:-1])).T

    @staticmethod
    def parser(item, site):
        page = requests.get(site, headers={'User-Agent': 'Mozilla/5.0'})
        tree = html.fromstring(page.content)
        week = list()

        if item in ("EONIA", "Libor 3M (USD)", "Euribor 3M"):
            last = tree.xpath('//table/tr[@class="tabledata1"]/td/text()')[1].replace("\xa0", "")

            for i in np.arange(1, 11, 2):
                week.insert(0, tree.xpath(
                    '//table/tr[@class="tabledata1"]/td/text() | //table/tr[@class="tabledata2"]/td/text()')[i].replace(
                    "\xa0", ""))
        else:
            arbre = tree.xpath('//table[@class="cr_dataTable"]/tbody')
            for t in arbre[0]:
                week.insert(1, t.text_content().split()[4].replace("'", ""))
            week = week[1:]
            last = tree.xpath('//span[@id="quote_val"]/text()')

        return week, last

    def run(self):
        self.load_data()
        self.get_df_change()
        return self.current, self.change

    def get_df_change(self):
        all_assets = pd.DataFrame()

        all_assets["Last Week"] = self.previous.iloc[:, 0]
        all_assets["This Week"] = self.current.iloc[:, -1]
        all_assets["As of Jan 1st"] = self.ytd_df

        rates = all_assets[self.type_spot.loc[all_assets.index] == 'rate'] / 100
        rates["Weekly Change"] = (rates["This Week"] - rates["As of Jan 1st"]) * 10000
        rates["YTD"] = (rates["This Week"] - rates["As of Jan 1st"]) * 10000

        other = all_assets[self.type_spot.loc[all_assets.index] == 'other']
        other["Weekly Change"] = other[["Last Week", "This Week"]].pct_change(axis=1)["This Week"]
        other["YTD"] = other[["This Week", "As of Jan 1st"]].pct_change(axis=1)["As of Jan 1st"]

        self.change = pd.concat((other, rates))
        self.change = self.change.reindex(['Last Week', 'This Week', 'Weekly Change', 'As of Jan 1st', 'YTD'], axis=1)
        self.change = self.change.reindex(self.source.index, axis=0)

    @staticmethod
    def offset_date_str(date, offset):
        return (dt.datetime.strptime(date, "%Y-%m-%d") + dt.timedelta(days=offset)).strftime("%Y-%m-%d")


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    from openpyxl import load_workbook
    import pandas as pd
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


if __name__ == '__main__':
    path = './Sources.xlsx'
    sources = pd.read_excel(path, sheet_name='sources', index_col=0)

    # load and format data
    prices, change = LoadDataWeekChange(sources['Source'], sources['Init'], sources['Type']).run()

    # write results in excel file
    writer = pd.ExcelWriter('./results.xlsx')
    prices.to_excel(writer, sheet_name='prices')
    change.to_excel(writer, sheet_name='changes')
    writer.save()
